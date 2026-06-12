import json
import os
import re
import sqlite3
import sys
from collections import defaultdict
from pathlib import Path

# Add scripts directory to path to import pitch_messages
sys.path.append("/home/prata/leads/scripts")
import pitch_messages

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

# Paths
db_path = "/home/prata/leads/data/leads.db"
serpapi_path = "/home/prata/leads/data/serpapi_leads.json"
wa_sent_path = "/home/prata/leads/data/wa_sent.json"
pitch_sheet_path = "/home/prata/leads/data/exports/pitch_sheet.md"
demos_dir = "/home/prata/leads/demos"
output_path = "/home/prata/leads/data/exports/Pitch_Book.xlsx"

# 1. Helper functions
def last_10_digits(phone):
    if not phone:
        return ""
    digits = re.sub(r'\D', '', str(phone))
    return digits[-10:] if len(digits) >= 10 else digits

def clean_name(name):
    if not name:
        return ""
    return re.sub(r'[^a-z0-9]', '', name.lower())

def get_niche_from_category(category, name):
    if not category:
        category = ""
    cat = category.lower()
    nm = name.lower()
    
    if "dent" in cat or "prosthodontist" in cat or "dent" in nm:
        return "dental"
    if "eyelash" in cat or "eyelash" in nm:
        return "eyelash"
    if "salon" in cat or "hair" in cat or "beauty" in cat or "makeup" in cat or "salon" in nm or "beauty" in nm:
        return "salon"
    if "barber" in cat or "barber" in nm:
        return "barber"
    if "cafe" in cat or "coffee" in cat or "cafe" in nm or "coffee" in nm:
        return "cafe"
    if "gym" in cat or "fitness" in cat or "gym" in nm or "fitness" in nm:
        return "gym"
    if "yoga" in cat or "pilates" in cat or "yoga" in nm or "pilates" in nm:
        return "yoga"
    if "skin" in cat or "dermat" in cat or "skin" in nm:
        return "skin"
    if "spa" in cat or "massage" in cat or "sauna" in cat or "spa" in nm:
        return "spa"
    if "physio" in cat or "therapy" in cat or "rehab" in cat or "physio" in nm:
        return "physio"
    if "bakery" in cat or "bake" in cat or "bakery" in nm or "bake" in nm:
        return "bakery"
    return "other"

def get_niche_whatsapp_message(niche, name, rating, address, demo_url):
    area = address.split(',')[0].strip() if address else "Mumbai"
    if len(area) > 15:
        area = area[:12] + "..."
        
    stars = "⭐" * min(5, round(rating)) if rating else ""
    rating_str = f"{rating}★ {stars}" if rating else "great rating"
    
    niche_lower = niche.lower()
    
    # Niche-specific templates
    if niche_lower == "dental":
        template = "Hi, I'm Kevin, a Mumbai web developer. Found {name} on Maps ({area}) with {rating_str}. Since you don't have a website, I built a quick preview site from your profile. Free to check out: {demo_url}"
    elif niche_lower == "salon":
        template = "Hi, I'm Kevin, a Mumbai web developer. Found your salon {name} on Maps ({area}) with {rating_str}. Since you don't have a website, I built a quick preview site from your profile. Free to check out: {demo_url}"
    elif niche_lower == "gym":
        template = "Hi, I'm Kevin, a Mumbai web developer. Saw {name} on Maps ({area}) with {rating_str}. Noticed no website, so I made a quick preview site showing your classes. Free to check out: {demo_url}"
    elif niche_lower == "barber":
        template = "Hi, I'm Kevin, a Mumbai web developer. Found {name} on Maps ({area}) with {rating_str}. Since you don't have a website, I built a quick preview site from your profile. Free to check out: {demo_url}"
    elif niche_lower == "skin":
        template = "Hi, I'm Kevin, a Mumbai web developer. Found your skin clinic {name} on Maps ({area}) with {rating_str}. Since you don't have a website, I built a quick preview site from your profile. Free to check out: {demo_url}"
    elif niche_lower == "cafe":
        template = "Hi, I'm Kevin, a Mumbai web developer. Found {name} on Maps ({area}) with {rating_str}. Noticed no website, so I made a quick preview site showing your menu. Free to check out: {demo_url}"
    elif niche_lower == "yoga":
        template = "Hi, I'm Kevin, a Mumbai web developer. Found your yoga studio {name} on Maps ({area}) with {rating_str}. Since you don't have a website, I built a quick preview site from your profile. Free to check out: {demo_url}"
    elif niche_lower == "physio":
        template = "Hi, I'm Kevin, a Mumbai web developer. Found your physio clinic {name} on Maps ({area}) with {rating_str}. Since you don't have a website, I built a quick preview site from your profile. Free to check out: {demo_url}"
    elif niche_lower == "bakery":
        template = "Hi, I'm Kevin, a Mumbai web developer. Found your bakery {name} on Maps ({area}) with {rating_str}. Since you don't have a website, I built a quick preview site from your profile. Free to check out: {demo_url}"
    elif niche_lower == "spa":
        template = "Hi, I'm Kevin, a Mumbai web developer. Found your spa {name} on Maps ({area}) with {rating_str}. Since you don't have a website, I built a quick preview site from your profile. Free to check out: {demo_url}"
    elif niche_lower == "eyelash":
        template = "Hi, I'm Kevin, a Mumbai web developer. Found your eyelash salon {name} on Maps ({area}) with {rating_str}. Since you don't have a website, I built a quick preview site from your profile. Free to check out: {demo_url}"
    else:
        template = "Hi, I'm Kevin, a Mumbai web developer. Found {name} on Maps ({area}) with {rating_str}. Since you don't have a website, I built a quick preview site from your profile. Free to check out: {demo_url}"
        
    # Truncate business name if needed to keep under 300 characters
    temp_no_name = template.format(name="", area=area, rating_str=rating_str, demo_url=demo_url)
    available_chars = 299 - len(temp_no_name)
    
    if len(name) > available_chars:
        # Truncate name to fit
        name = name[:available_chars - 3] + "..."
        
    return template.format(name=name, area=area, rating_str=rating_str, demo_url=demo_url)

# Load database leads
conn = sqlite3.connect(db_path)
conn.row_factory = sqlite3.Row
cursor = conn.cursor()
cursor.execute("SELECT * FROM leads;")
db_rows = [dict(r) for r in cursor.fetchall()]
conn.close()

db_by_phone10 = {}
db_by_name = {}
for r in db_rows:
    p10 = last_10_digits(r['phone'])
    if p10:
        db_by_phone10[p10] = r
    cname = clean_name(r['name'])
    if cname:
        db_by_name[cname] = r

# Parse pitch_sheet.md for details
pitch_by_phone10 = {}
pitch_by_name = {}
if os.path.exists(pitch_sheet_path):
    with open(pitch_sheet_path, 'r') as f:
        pitch_content = f.read()
    
    # Split by level 3 headers
    full_scripts_part = pitch_content.split("## 📞 Full Pitch Scripts")[-1]
    individual_sections = re.split(r'\n###\s+', full_scripts_part)
    for sec in individual_sections[1:]:
        lines = sec.strip().split('\n')
        if not lines:
            continue
        header = lines[0]
        match = re.match(r'^\d+\.\s+(.+)$', header)
        if not match:
            continue
        biz_name = match.group(1).strip()
        
        # Extract phone, rating, address
        phone_rating_line = ""
        address_line = ""
        for line in lines[1:]:
            if "**Phone:**" in line:
                phone_rating_line = line
            elif "**Address:**" in line:
                address_line = line
                
        phone = ""
        rating = None
        reviews = 0
        if phone_rating_line:
            p_match = re.search(r'\*\*Phone:\*\*\s*(.*?)\s*\|', phone_rating_line)
            if p_match:
                phone = p_match.group(1).strip()
            r_match = re.search(r'\*\*Rating:\*\*\s*([\d\.]+)⭐', phone_rating_line)
            if r_match:
                rating = float(r_match.group(1))
            rev_match = re.search(r'\(([\d,]+)\s+reviews\)', phone_rating_line)
            if rev_match:
                reviews = int(rev_match.group(1).replace(",", ""))
                
        address = ""
        if address_line:
            a_match = re.search(r'\*\*Address:\*\*\s*(.*)$', address_line)
            if a_match:
                address = a_match.group(1).strip()
                
        p10 = last_10_digits(phone)
        info = {
            "name": biz_name,
            "phone": phone,
            "rating": rating,
            "reviews": reviews,
            "address": address
        }
        if p10:
            pitch_by_phone10[p10] = info
        cname = clean_name(biz_name)
        if cname:
            pitch_by_name[cname] = info

# Load JSON data
with open(serpapi_path, 'r') as f:
    serpapi_leads = json.load(f)

with open(wa_sent_path, 'r') as f:
    wa_sent = json.load(f)

# Get demo folders
demo_folders = os.listdir(demos_dir) if os.path.exists(demos_dir) else []
folder_map = {clean_name(f): f for f in demo_folders}

def get_demo_url(name, phone):
    # Try direct name slug match
    cname = clean_name(name)
    if cname in folder_map:
        return f"https://kevinpratap.github.io/lead-demos/{folder_map[cname]}/"
    
    # Try phone resolved name
    p10 = last_10_digits(phone)
    resolved_name = ""
    if p10 in db_by_phone10:
        resolved_name = db_by_phone10[p10]['name']
    elif p10 in pitch_by_phone10:
        resolved_name = pitch_by_phone10[p10]['name']
        
    if resolved_name:
        rcname = clean_name(resolved_name)
        if rcname in folder_map:
            return f"https://kevinpratap.github.io/lead-demos/{folder_map[rcname]}/"
            
    # Try generic slug check
    slug = pitch_messages._slug({"name": name})
    if slug in demo_folders:
        return f"https://kevinpratap.github.io/lead-demos/{slug}/"
        
    return None

# Build lookup of WA sent messages by phone to merge duplicates
# And get the earliest sent_at date
wa_sent_by_phone10 = defaultdict(list)
for item in wa_sent:
    p10 = last_10_digits(item.get('phone'))
    if p10:
        wa_sent_by_phone10[p10].append(item)

# 2. Compile unique leads for Sheet 1
all_leads_dict = {}

# Process existing leads from WA sent
for p10, items in wa_sent_by_phone10.items():
    # Find the earliest message and check if any succeeded
    success = "No"
    sent_dates = []
    for it in items:
        if it.get('success'):
            success = "Yes"
        if it.get('sent_at'):
            sent_dates.append(it.get('sent_at'))
            
    earliest_sent = min(sent_dates) if sent_dates else None
    
    # Resolve canonical info
    db_info = db_by_phone10.get(p10)
    pitch_info = pitch_by_phone10.get(p10)
    
    # Fallback cascade
    name = ""
    if db_info:
        name = db_info['name']
    elif pitch_info:
        name = pitch_info['name']
    else:
        # Get first non-phone name from items
        for it in items:
            if it.get('lead') and not re.match(r'^\d+$', it.get('lead')):
                name = it.get('lead')
                break
        if not name:
            name = items[0].get('lead') # fallback
            
    phone = db_info['phone'] if db_info and db_info.get('phone') else (pitch_info['phone'] if pitch_info and pitch_info.get('phone') else items[0].get('phone'))
    address = db_info['address'] if db_info and db_info.get('address') else (pitch_info['address'] if pitch_info and pitch_info.get('address') else "")
    rating = db_info['rating'] if db_info and db_info.get('rating') is not None else (pitch_info['rating'] if pitch_info and pitch_info.get('rating') is not None else None)
    reviews = db_info['reviews'] if db_info and db_info.get('reviews') is not None else (pitch_info['reviews'] if pitch_info and pitch_info.get('reviews') is not None else 0)
    website = db_info['website'] if db_info and db_info.get('website') else ""
    category = db_info['category'] if db_info and db_info.get('category') else ""
    
    niche = get_niche_from_category(category, name)
    
    # Try to extract demo url from sent messages
    demo_url = None
    for it in items:
        msg = it.get('message', '')
        url_match = re.search(r'https://kevinpratap\.github\.io/lead-demos/([a-z0-9\-]+)/?', msg)
        if url_match:
            demo_url = url_match.group(0)
            break
            
    if not demo_url:
        demo_url = get_demo_url(name, phone)
        
    all_leads_dict[p10] = {
        "name": name,
        "niche": niche,
        "phone": phone,
        "address": address,
        "rating": rating,
        "reviews": reviews,
        "website": website,
        "demo_url": demo_url,
        "wa_sent": "Yes",
        "sent_date": earliest_sent,
        "source": "previous_wa"
    }

# Process new leads from SerpAPI
for l in serpapi_leads:
    name = l.get('name')
    phones = l.get('phones', [])
    phone = phones[0] if phones else ""
    p10 = last_10_digits(phone)
    
    # Check if this phone was already processed in WA sent
    if p10 and p10 in all_leads_dict:
        # Already processed, skip
        continue
        
    address = l.get('address') or ""
    rating = l.get('rating')
    website = l.get('website') or ""
    niche = l.get('niche') or ""
    source = "serpapi_organic" if l.get('source') == 'organic' else "serpapi_local"
    
    demo_url = get_demo_url(name, phone)
    
    lead_key = p10 if p10 else f"no_phone_{clean_name(name)}"
    
    # Try to resolve reviews from DB if matched by name
    cname = clean_name(name)
    reviews = 0
    if cname in db_by_name:
        reviews = db_by_name[cname].get('reviews', 0)
    elif cname in pitch_by_name:
        reviews = pitch_by_name[cname].get('reviews', 0)
        
    all_leads_dict[lead_key] = {
        "name": name,
        "niche": niche,
        "phone": phone,
        "address": address,
        "rating": rating,
        "reviews": reviews,
        "website": website,
        "demo_url": demo_url,
        "wa_sent": "No",
        "sent_date": None,
        "source": source
    }

all_leads = list(all_leads_dict.values())
print(f"Total compiled unique leads: {len(all_leads)}")

# 3. Create Excel workbook
wb = openpyxl.Workbook()

# Sheet 1: "All Leads Master"
ws1 = wb.active
ws1.title = "All Leads Master"
ws1_headers = [
    "Business Name", "Niche", "Phone Number", "Address", "Rating",
    "Website URL", "Demo Site URL", "WhatsApp Sent", "Sent Date", "Source"
]
ws1.append(ws1_headers)

for l in all_leads:
    ws1.append([
        l["name"], l["niche"], l["phone"], l["address"], l["rating"],
        l["website"], l["demo_url"], l["wa_sent"], l["sent_date"], l["source"]
    ])

# Sheet 2: "Ready to Pitch"
ws2 = wb.create_sheet(title="Ready to Pitch")
ws2_headers = ws1_headers + ["Pitch Script", "Call Script", "Priority Score"]
ws2.append(ws2_headers)

# Filter and calculate scores
ready_to_pitch = []
for l in all_leads:
    if l["phone"] and l["wa_sent"] == "No":
        # Calculate priority score: Rating * 10 + Reviews * 0.05
        rating_val = l["rating"] if l["rating"] is not None else 0.0
        reviews_val = l.get("reviews") or 0
        priority_score = rating_val * 10 + reviews_val * 0.05
        
        # Generate expected/dummy Demo URL for length checking and scripts
        lead_dict = {
            "name": l["name"],
            "phone": l["phone"],
            "rating": l["rating"],
            "address": l["address"],
            "category": l["niche"]
        }
        slug = pitch_messages._slug(lead_dict)
        expected_demo_url = f"https://kevinpratap.github.io/lead-demos/{slug}/"
        
        # Generate pitch script using our length-constrained, niche-specific builder
        pitch_script = get_niche_whatsapp_message(l["niche"], l["name"], l["rating"], l["address"], "{demo_url}")
        
        # Call Script
        call_script_text = pitch_messages.call_script(lead_dict)
        
        ready_to_pitch.append({
            "lead": l,
            "pitch_script": pitch_script,
            "call_script": call_script_text,
            "priority_score": priority_score,
            "expected_demo_url": expected_demo_url
        })

# Sort by priority score descending
ready_to_pitch.sort(key=lambda x: x["priority_score"], reverse=True)

for r in ready_to_pitch:
    l = r["lead"]
    ws2.append([
        l["name"], l["niche"], l["phone"], l["address"], l["rating"],
        l["website"], l["demo_url"], l["wa_sent"], l["sent_date"], l["source"],
        r["pitch_script"], r["call_script"], r["priority_score"]
    ])

# Sheet 3: "WhatsApp Scripts"
ws3 = wb.create_sheet(title="WhatsApp Scripts")
ws3_headers = [
    "Business Name", "Phone", "Niche", "WhatsApp Message", "Call Script", "Demo URL"
]
ws3.append(ws3_headers)

for r in ready_to_pitch:
    l = r["lead"]
    lead_dict = {
        "name": l["name"],
        "phone": l["phone"],
        "rating": l["rating"],
        "address": l["address"],
        "category": l["niche"]
    }
    
    # Generate short call script: opening and hook
    addr_part = l["address"].split(',')[0].strip() if l["address"] else "great reviews"
    rating_val = f"{l['rating']:.1f}" if l['rating'] is not None else "—"
    short_call = (
        f"Hi, is this {l['name']}? Can I speak with the owner/doctor?\n\n"
        f"I'm Kevin, web developer. Found you on Google Maps — {rating_val}★, {addr_part}. "
        f"Noticed you don't have a website. I actually built a preview website for {l['name']} already — "
        f"can I send you the link on WhatsApp to show you? Takes 10 seconds to look."
    )
    
    # WhatsApp Message with {demo_url} placeholder
    ws3.append([
        l["name"], l["phone"], l["niche"], r["pitch_script"], short_call, r["expected_demo_url"]
    ])

# Sheet 4: "Sent Log"
ws4 = wb.create_sheet(title="Sent Log")
ws4_headers = ["Business Name", "Phone", "Sent Date/Time", "Success", "Demo URL used"]
ws4.append(ws4_headers)

# Sort wa_sent by sent_at descending
wa_sent_sorted = sorted(wa_sent, key=lambda x: x.get('sent_at', ''), reverse=True)

for s in wa_sent_sorted:
    phone = s.get('phone', '')
    p10 = last_10_digits(phone)
    
    # Resolve business name
    db_info = db_by_phone10.get(p10)
    pitch_info = pitch_by_phone10.get(p10)
    name = ""
    if db_info:
        name = db_info['name']
    elif pitch_info:
        name = pitch_info['name']
    else:
        name = s.get('lead')
        
    success_val = "Yes" if s.get('success') else "No"
    
    # Extract Demo URL from message
    demo_url_used = ""
    msg = s.get('message', '')
    url_match = re.search(r'https://kevinpratap\.github\.io/lead-demos/([a-z0-9\-]+)/?', msg)
    if url_match:
        demo_url_used = url_match.group(0)
    else:
        # Fallback to get_demo_url
        demo_url_used = get_demo_url(name, phone) or ""
        
    ws4.append([
        name, phone, s.get('sent_at'), success_val, demo_url_used
    ])

# Sheet 5: "Summary Dashboard"
ws5 = wb.create_sheet(title="Summary Dashboard")
ws5.views.sheetView[0].showGridLines = True

# Styling helpers
title_font = Font(name='Segoe UI', size=16, bold=True, color='FFFFFF')
title_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
kpi_fill = PatternFill(start_color='EAECEE', end_color='EAECEE', fill_type='solid')
bold_font = Font(name='Segoe UI', size=11, bold=True)
regular_font = Font(name='Segoe UI', size=11)
thin_border = Border(
    left=Side(style='thin', color='D3D3D3'),
    right=Side(style='thin', color='D3D3D3'),
    top=Side(style='thin', color='D3D3D3'),
    bottom=Side(style='thin', color='D3D3D3')
)

# Set grid line visibility for all sheets explicitly
for sheet in wb.worksheets:
    sheet.views.sheetView[0].showGridLines = True

# Title row for Dashboard
ws5.merge_cells('A1:E1')
title_cell = ws5['A1']
title_cell.value = "  LEAD GENERATION OUTREACH DASHBOARD"
title_cell.font = title_font
title_cell.fill = title_fill
title_cell.alignment = Alignment(vertical='center')
ws5.row_dimensions[1].height = 40

# Key Performance Indicators Table
ws5['A3'] = "Key Performance Indicators"
ws5['A3'].font = Font(name='Segoe UI', size=12, bold=True, color='1F4E79')
ws5.merge_cells('A3:B3')

ws5['A4'] = "Metric"
ws5['B4'] = "Value"
ws5['A4'].font = header_font
ws5['A4'].fill = header_fill
ws5['A4'].alignment = Alignment(horizontal='center')
ws5['B4'].font = header_font
ws5['B4'].fill = header_fill
ws5['B4'].alignment = Alignment(horizontal='center')

kpis = [
    ("Total Unique Leads", f"=COUNTA('All Leads Master'!A2:A{ws1.max_row})"),
    ("Leads with Phone Numbers", f"=COUNTA('All Leads Master'!C2:C{ws1.max_row}) - COUNTBLANK('All Leads Master'!C2:C{ws1.max_row})"),
    ("WhatsApp Messages Sent", f"=COUNTIF('All Leads Master'!H2:H{ws1.max_row}, \"Yes\")"),
    ("Remaining to Pitch", f"=COUNTA('Ready to Pitch'!A2:A{ws2.max_row})"),
    ("Demo Sites Generated", f"=COUNTA('All Leads Master'!G2:G{ws1.max_row}) - COUNTBLANK('All Leads Master'!G2:G{ws1.max_row})")
]

for idx, (metric, formula) in enumerate(kpis, start=5):
    ws5.cell(row=idx, column=1, value=metric).font = bold_font
    ws5.cell(row=idx, column=1).border = thin_border
    
    val_cell = ws5.cell(row=idx, column=2, value=formula)
    val_cell.font = regular_font
    val_cell.alignment = Alignment(horizontal='right')
    val_cell.border = thin_border
    val_cell.fill = kpi_fill

# Leads by Niche Table
ws5['D3'] = "Leads by Target Niche"
ws5['D3'].font = Font(name='Segoe UI', size=12, bold=True, color='1F4E79')
ws5.merge_cells('D3:E3')

ws5['D4'] = "Niche"
ws5['E4'] = "Lead Count"
ws5['D4'].font = header_font
ws5['D4'].fill = header_fill
ws5['D4'].alignment = Alignment(horizontal='center')
ws5['E4'].font = header_font
ws5['E4'].fill = header_fill
ws5['E4'].alignment = Alignment(horizontal='center')

niches = [
    "dental", "salon", "gym", "barber", "skin", "cafe", 
    "yoga", "physio", "bakery", "spa", "eyelash"
]

for idx, niche in enumerate(niches, start=5):
    ws5.cell(row=idx, column=4, value=niche.capitalize()).font = regular_font
    ws5.cell(row=idx, column=4).border = thin_border
    
    formula = f"=COUNTIF('All Leads Master'!B2:B{ws1.max_row}, \"{niche}\")"
    val_cell = ws5.cell(row=idx, column=5, value=formula)
    val_cell.font = regular_font
    val_cell.alignment = Alignment(horizontal='right')
    val_cell.border = thin_border

# Add Total row for Niches
ws5.cell(row=16, column=4, value="Total").font = bold_font
ws5.cell(row=16, column=4).border = thin_border
total_niche_cell = ws5.cell(row=16, column=5, value=f"=SUM(E5:E15)")
total_niche_cell.font = bold_font
total_niche_cell.alignment = Alignment(horizontal='right')
total_niche_cell.border = thin_border

# Auto-adjust column widths for Dashboard
ws5.column_dimensions['A'].width = 30
ws5.column_dimensions['B'].width = 15
ws5.column_dimensions['C'].width = 5
ws5.column_dimensions['D'].width = 25
ws5.column_dimensions['E'].width = 18

# 4. Apply formatting to Sheets 1-4
def format_data_sheet(ws, is_pitch_sheet=False):
    h_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
    h_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid') # Navy Blue
    h_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    zebra_fill = PatternFill(start_color='F2F4F7', end_color='F2F4F7', fill_type='solid')
    reg_font = Font(name='Segoe UI', size=10)
    
    t_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )
    
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = h_font
        cell.fill = h_fill
        cell.alignment = h_align
        cell.border = t_border
        
    ws.row_dimensions[1].height = 28
        
    for row_idx in range(2, ws.max_row + 1):
        is_even = row_idx % 2 == 0
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = reg_font
            cell.border = t_border
            if is_even:
                cell.fill = zebra_fill
                
            # Alignments
            if col_idx in [2, 5, 8, 10]:  # Niche, Rating, WA Sent, Source
                cell.alignment = Alignment(horizontal='center', vertical='top')
            elif col_idx in [11, 12] and is_pitch_sheet:  # Pitch Script, Call Script (long texts)
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal='left', vertical='top')
                
    ws.freeze_panes = 'A2'
    
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        col_idx = col[0].column
        
        if is_pitch_sheet and col_idx in [11, 12]:  # Pitch Script, Call Script
            ws.column_dimensions[col_letter].width = 50
            continue
            
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > 100:
                val_str = val_str[:100]
            max_len = max(max_len, len(val_str))
            
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

format_data_sheet(ws1)
format_data_sheet(ws2, is_pitch_sheet=True)
format_data_sheet(ws3)
format_data_sheet(ws4)

# Apply conditional formatting for Priority Score on Sheet 2 (Column M)
if len(ready_to_pitch) > 0:
    color_scale = ColorScaleRule(
        start_type='num', start_value=0.0, start_color='F8CBAD', # light red
        mid_type='num', mid_value=45.0, mid_color='FFF2CC',      # light yellow
        end_type='num', end_value=90.0, end_color='C6E0B4'       # light green
    )
    ws2.conditional_formatting.add(f'M2:M{ws2.max_row}', color_scale)

# Save workbook
os.makedirs(os.path.dirname(output_path), exist_ok=True)
wb.save(output_path)
print(f"\nExcel workbook saved successfully to {output_path}!")

# Count calculations for terminal print summary
print("\n" + "="*40)
print("             TASK SUMMARY")
print("="*40)
print(f"Total Unique Leads:            {len(all_leads)}")
print(f"Leads with Phone Numbers:      {len([l for l in all_leads if l['phone']])}")
print(f"WhatsApp Messages Sent:        {len([l for l in all_leads if l['wa_sent'] == 'Yes'])}")
print(f"Remaining to Pitch:            {len(ready_to_pitch)}")
print(f"Demo Sites Generated:          {len([l for l in all_leads if l['demo_url']])}")
print("="*40)
print("Leads by Niche:")
niche_counts = defaultdict(int)
for l in all_leads:
    niche_counts[l['niche']] += 1
for niche in sorted(niches):
    print(f"  - {niche.capitalize()}: " + " "*(15-len(niche)) + f"{niche_counts[niche]}")
print("="*40)
